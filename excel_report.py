"""將投資報告輸出為 Excel（多工作表），對應 report.py 的固定格式八大區塊。"""

from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from portfolio import (
    load_portfolio,
    enrich_with_market_data,
    build_report_context,
    weight,
    cost_weight,
    format_price_date,
    industry_or_sector_label_zh,
    market_intel_highlight,
    MAX_CARRYOVER_MONTHS,
)
from report import DISCLAIMER

EXCEL_REPORT_PATH = "投資報告.xlsx"

HEADER_FONT = Font(bold=True, size=16)
DETAIL_FONT_SIZE = 14


def _write_header(sheet, row, headers):
    for col, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(sheet, widths):
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width


def _rescale_fonts(wb):
    """全報表明細字體統一調成 14、標題（欄位表頭，HEADER_FONT）維持 16。
    走訪每個分頁已使用的儲存格，跳過已經是標題大小（16）的儲存格，其餘一律改成 14，
    同時保留原本的粗體／顏色設定（例如損益紅漲綠跌、合理買點的藍色粗體、大盤漲跌配色）。
    """
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                font = cell.font
                if font.size == 16:
                    continue
                cell.font = Font(bold=font.bold, italic=font.italic, underline=font.underline, color=font.color, size=DETAIL_FONT_SIZE)


def _sheet_overview(wb, holdings, ctx):
    sheet = wb.active
    sheet.title = "投資組合總覽"

    sheet["A1"] = "投資組合總覽"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"更新時間：{datetime.now().strftime('%Y-%m-%d %H:%M')}"

    sheet["A4"] = "總成本"
    sheet["B4"] = ctx["total_cost"]
    sheet["A5"] = "總市值"
    sheet["B5"] = ctx["total_value"]
    total_pnl_color = "C62828" if (ctx["total_pnl"] or 0) >= 0 else "2E7D32"  # 台股慣例：紅漲綠跌
    total_pnl_font = Font(color=total_pnl_color)
    sheet["A6"] = "總報酬"
    sheet["B6"] = ctx["total_pnl"]
    sheet["B6"].font = total_pnl_font
    sheet["A7"] = "總報酬率"
    sheet["B7"] = ctx["total_pnl_pct"]
    sheet["B7"].font = total_pnl_font
    sheet["B7"].number_format = "0.0%"

    idx = ctx.get("market_index")
    if idx and idx.get("close") is not None:
        sheet["A8"] = f"大盤{idx['name']}（{idx.get('date', '')}）"
        sheet["B8"] = idx["close"]
        if idx.get("change_points") is not None:
            sheet["C8"] = idx["change_points"]
            sheet["D8"] = idx["change_pct"] / 100 if idx.get("change_pct") is not None else None
            sheet["D8"].number_format = "0.00%"

    headers = ["代號", "名稱", "股數", "平均成本", "現價", "市值", "損益", "損益率", "市值佔比", "成本佔比"]
    start_row = 9
    _write_header(sheet, start_row, headers)
    sheet.freeze_panes = f"A{start_row + 1}"

    for i, h in enumerate(holdings, start=start_row + 1):
        w = weight(h, ctx["total_value"])
        cw = cost_weight(h, ctx["total_cost"])
        sheet.cell(i, 1, h.code)
        sheet.cell(i, 2, h.name)
        sheet.cell(i, 3, h.shares)
        sheet.cell(i, 4, h.avg_cost)
        sheet.cell(i, 5, h.price)
        sheet.cell(i, 6, h.market_value)
        pnl_color = "C62828" if (h.pnl or 0) >= 0 else "2E7D32"  # 台股慣例：紅漲綠跌
        pnl_font = Font(color=pnl_color)
        c = sheet.cell(i, 7, h.pnl)
        c.font = pnl_font
        c = sheet.cell(i, 8, h.pnl_pct)
        c.font = pnl_font
        c.number_format = "0.0%"
        c = sheet.cell(i, 9, w)
        c.number_format = "0.0%"
        c = sheet.cell(i, 10, cw)
        c.number_format = "0.0%"

    _autofit(sheet, [10, 20, 10, 10, 10, 12, 12, 10, 10, 10])


def _price_date_header_label(holdings) -> str | None:
    """取眾數（最常見的資料日期）當欄位標題的日期標籤，避免少數幾檔抓到不同日期就整欄跟著亂標。"""
    dates = [h.price_date for h in holdings if h.price_date]
    if not dates:
        return None
    most_common_raw = Counter(dates).most_common(1)[0][0]
    return format_price_date(most_common_raw)


def _sheet_fundamentals(wb, holdings):
    sheet = wb.create_sheet("個股分析")
    date_label = _price_date_header_label(holdings)
    price_header = f"目前股價({date_label})" if date_label else "目前股價"
    headers = ["代號", "名稱", "產業", "本益比(TTM)", "本益比(預估)", "EPS(TTM)", "EPS(預估)", price_header, "法人目標價", "法人評等", "殖利率"]
    _write_header(sheet, 1, headers)
    sheet.freeze_panes = "A2"
    for i, h in enumerate(holdings, start=2):
        sheet.cell(i, 1, h.code)
        sheet.cell(i, 2, h.name)
        sheet.cell(i, 3, industry_or_sector_label_zh(h))
        sheet.cell(i, 4, h.trailing_pe)
        sheet.cell(i, 5, h.forward_pe)
        sheet.cell(i, 6, h.trailing_eps)
        sheet.cell(i, 7, h.forward_eps)
        sheet.cell(i, 8, h.price)
        sheet.cell(i, 9, h.target_mean_price)
        sheet.cell(i, 10, h.recommendation or "N/A")
        c = sheet.cell(i, 11, h.dividend_yield)
        if h.dividend_yield:
            c.number_format = "0.00%"
    _autofit(sheet, [10, 20, 16, 12, 12, 10, 10, 20, 12, 12, 10])


RANGE_HIGHLIGHT_FONT = Font(bold=True, color="1864AB")  # 跟太太盤中觀察信件的股票代號同一套醒目藍

# portfolio.buy_point_data() 的 price_position 是給 Markdown/PDF/網頁版看的完整句子；
# Excel 表格欄位窄，只在這裡另外做簡述版，不動共用的資料函式（避免影響其他三種輸出格式）。
_PRICE_POSITION_BRIEF = {
    "現價已落在偏低區間，屬於合理承接位置": "偏低，可承接",
    "現價高於偏低承接區間，建議等待拉回或分批布局": "偏高，待拉回或分批",
    "尚無近一年價格區間資料，暫無法估算買點。": "無資料",
}


def _brief_price_position(text: str | None) -> str:
    if not text:
        return "N/A"
    return _PRICE_POSITION_BRIEF.get(text, text)


def _sheet_buy_points(wb, buy_points_data: list[dict]):
    sheet = wb.create_sheet("合理買點")
    headers = ["代號", "名稱", "近一年區間", "建議承接區間", "現價位置", "法人目標價", "上漲空間"]
    _write_header(sheet, 1, headers)
    sheet.freeze_panes = "A2"
    for i, bp in enumerate(buy_points_data, start=2):
        sheet.cell(i, 1, bp["code"])
        sheet.cell(i, 2, bp["name"])
        sheet.cell(i, 3, bp.get("week52_range") or "N/A").font = RANGE_HIGHLIGHT_FONT
        sheet.cell(i, 4, bp.get("suggested_zone") or "N/A").font = RANGE_HIGHLIGHT_FONT
        sheet.cell(i, 5, _brief_price_position(bp.get("price_position")))
        sheet.cell(i, 5).alignment = Alignment(wrap_text=True)
        sheet.cell(i, 6, bp.get("target_price"))
        c = sheet.cell(i, 7, bp.get("upside_pct"))
        c.number_format = "0.0%"
    _autofit(sheet, [10, 20, 18, 18, 40, 14, 12])


def _sheet_monthly_allocation(wb, data: dict):
    """需求 6：優先購買清單＋資金分配改成表格分頁，總預算說明／尚未分配結轉／排除加碼說明
    比照既有 `_sheet_sector_allocation()` 的「表格+文字說明」混合寫法，寫在表格上方／下方。"""
    sheet = wb.create_sheet("每月投資建議")

    if data.get("no_value_note"):
        sheet.cell(1, 1, data["no_value_note"])
        _autofit(sheet, [60])
        return

    amount = data["amount"]
    carryover = data["carryover"]
    total_budget = data["total_budget"]
    budget_line = (
        f"本月可投資金額：新台幣 {total_budget:,.0f} 元"
        + (f"（月定期定額 {amount:,.0f} 元＋現金流結餘 {carryover:,.0f} 元，結轉上限 {data['carryover_cap']:,.0f} 元）" if carryover else f"（月定期定額 {amount:,.0f} 元）")
    )
    sheet.cell(1, 1, budget_line).font = Font(bold=True, size=12)
    sheet.freeze_panes = "A2"
    row = 3

    if data.get("no_candidates_note"):
        sheet.cell(row, 1, data["no_candidates_note"])
        row += 1
        for note in data.get("excluded_notes") or []:
            sheet.cell(row, 1, note)
            sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
            row += 1
        _autofit(sheet, [100])
        return

    sheet.cell(row, 1, "優先購買清單（依 Investment Score／長期持股信心排序）").font = HEADER_FONT
    row += 1
    _write_header(sheet, row, ["順位", "代號", "名稱", "Investment Score", "長期信心"])
    row += 1
    for p in data["priority_list"]:
        sheet.cell(row, 1, p["rank"])
        sheet.cell(row, 2, p["code"])
        sheet.cell(row, 3, p["name"])
        sheet.cell(row, 4, round(p["investment_score"], 1) if p.get("investment_score") is not None else None)
        sheet.cell(row, 5, p.get("conviction_stars", "N/A"))
        row += 1

    row += 1
    sheet.cell(row, 1, "本月資金分配（集中在優先順序最前面的標的，不是每檔都買一點；資金用完或達上限為止）").font = HEADER_FONT
    row += 1
    _write_header(sheet, row, ["代號", "名稱", "投入金額", "買進階段", "本輪比例", "約可買股數", "備註"])
    row += 1
    for a in data["allocations"]:
        sheet.cell(row, 1, a["code"])
        sheet.cell(row, 2, a["name"])
        sheet.cell(row, 3, a["amount"])
        sheet.cell(row, 4, a["tier_label"])
        c = sheet.cell(row, 5, a["tier_pct"])
        c.number_format = "0%"
        sheet.cell(row, 6, round(a["shares"], 1))
        sheet.cell(row, 7, a.get("note") or "")
        row += 1

    row += 1
    if data.get("remaining") is not None:
        sheet.cell(row, 1, f"尚未分配：{data['remaining']:,.0f} 元，將結轉至下月（結轉上限：{MAX_CARRYOVER_MONTHS} 個月投入金額 {data['carryover_cap']:,.0f} 元）。")
        row += 1
    if data.get("no_allocation_note"):
        sheet.cell(row, 1, data["no_allocation_note"])
        row += 1
    for note in data.get("excluded_notes") or []:
        sheet.cell(row, 1, note)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1

    _autofit(sheet, [10, 20, 14, 20, 12, 14, 40])


def _sheet_sector_allocation(wb, ctx):
    sheet = wb.create_sheet("產業配置建議")
    _write_header(sheet, 1, ["產業別", "佔投組市值比重"])
    sheet.freeze_panes = "A2"
    row = 2
    for sector, w in ctx["sector_allocation"].items():
        sheet.cell(row, 1, sector)
        c = sheet.cell(row, 2, w)
        c.number_format = "0.0%"
        row += 1

    row += 1
    sheet.cell(row, 1, "配置建議").font = HEADER_FONT
    row += 1
    for advice in ctx["sector_advice"]:
        sheet.cell(row, 1, advice)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1

    _autofit(sheet, [90, 16])


def _sheet_risk_rebalance(wb, ctx):
    sheet = wb.create_sheet("風險評估與再平衡")
    sheet.cell(1, 1, "風險評估").font = HEADER_FONT
    sheet.freeze_panes = "A2"
    row = 2
    for r in ctx["risk"]:
        sheet.cell(row, 1, r)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    sheet.cell(row, 1, "再平衡建議").font = HEADER_FONT
    row += 1
    for r in ctx["rebalance"]:
        sheet.cell(row, 1, r)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1

    _autofit(sheet, [100])


def _sheet_ai_scoring(wb, holdings, ai_scoring: dict):
    sheet = wb.create_sheet("AI綜合評分")
    headers = ["代號", "名稱", "訊號", "說明", "投資評分", "風險評分", "資料覆蓋率", "法人目標價", "合理價", "關鍵理由"]
    _write_header(sheet, 1, headers)
    sheet.freeze_panes = "A2"
    for i, h in enumerate(holdings, start=2):
        r = ai_scoring.get(h.code) or {}
        sheet.cell(i, 1, h.code)
        sheet.cell(i, 2, h.name)
        sheet.cell(i, 3, r.get("signal"))
        sheet.cell(i, 4, r.get("label"))
        sheet.cell(i, 5, round(r["investment_score"], 1) if r.get("investment_score") is not None else None)
        sheet.cell(i, 6, round(r["risk_score"], 1) if r.get("risk_score") is not None else None)
        c = sheet.cell(i, 7, r.get("data_coverage"))
        c.number_format = "0%"
        sheet.cell(i, 8, r.get("target_price"))
        sheet.cell(i, 9, r.get("fair_value"))
        sheet.cell(i, 10, "；".join(r.get("reasons") or []))
        sheet.cell(i, 10).alignment = Alignment(wrap_text=True)

    _autofit(sheet, [10, 20, 8, 50, 16, 12, 12, 14, 14, 60])


def _sheet_candidate_watchlist(wb, candidates: list[dict], sheet_name: str = "潛力新標的觀察"):
    sheet = wb.create_sheet(sheet_name)
    headers = ["價位帶", "代號", "名稱", "股價", "信心星等", "Investment Score", "Risk Score", "資料覆蓋率", "法人目標價", "關鍵理由"]
    _write_header(sheet, 1, headers)
    sheet.freeze_panes = "A2"
    tier_label = {"high": "高價股", "low": "低價股"}
    for i, r in enumerate(candidates, start=2):
        sheet.cell(i, 1, tier_label.get(r.get("price_tier"), r.get("price_tier")))
        sheet.cell(i, 2, r.get("code"))
        sheet.cell(i, 3, r.get("name"))
        sheet.cell(i, 4, r.get("price"))
        sheet.cell(i, 5, r.get("conviction_stars"))
        sheet.cell(i, 6, round(r["investment_score"], 1) if r.get("investment_score") is not None else None)
        sheet.cell(i, 7, round(r["risk_score"], 1) if r.get("risk_score") is not None else None)
        c = sheet.cell(i, 8, r.get("data_coverage"))
        c.number_format = "0%"
        sheet.cell(i, 9, r.get("target_price"))
        sheet.cell(i, 10, "；".join(r.get("reasons") or []))
        sheet.cell(i, 10).alignment = Alignment(wrap_text=True)

    _autofit(sheet, [10, 10, 20, 10, 12, 16, 12, 12, 14, 60])


def _sheet_market_intel(wb, holdings, market_intel: dict):
    sheet = wb.create_sheet("市場情報")
    headers = [
        "代號", "名稱", "重點摘要", "月營收年增率", "最新季", "季營收(仟元)", "EPS(元)", "今日重大訊息則數", "下一場法說會",
        "三大法人買賣超(股)", "外資買賣超", "投信買賣超", "自營商買賣超", "相關新聞1", "相關新聞2", "相關新聞3",
    ]
    _write_header(sheet, 1, headers)
    sheet.freeze_panes = "A2"
    for i, h in enumerate(holdings, start=2):
        info = market_intel.get(h.code) or {}
        rev = info.get("monthly_revenue") or {}
        inc = info.get("income_statement") or {}
        news = info.get("general_news") or []
        material_news = info.get("material_news") or []
        conf = info.get("investor_conference")
        flow = info.get("institutional_flow") or {}

        sheet.cell(i, 1, h.code)
        sheet.cell(i, 2, h.name)
        sheet.cell(i, 3, market_intel_highlight(info))
        sheet.cell(i, 3).alignment = Alignment(wrap_text=True)

        yoy = rev.get("營業收入-去年同月增減(%)")
        c = sheet.cell(i, 4, float(yoy) / 100 if yoy not in (None, "") else None)
        c.number_format = "0.0%"

        sheet.cell(i, 5, inc.get("quarter"))
        sheet.cell(i, 6, float(inc["revenue"]) if inc.get("revenue") is not None else None)
        sheet.cell(i, 7, float(inc["eps"]) if inc.get("eps") not in (None, "") else None)
        sheet.cell(i, 8, len(material_news))
        sheet.cell(i, 9, f"{conf['date']} {conf['time']}／{conf['location']}" if conf else None)
        sheet.cell(i, 10, flow.get("total_net"))
        sheet.cell(i, 11, flow.get("foreign_net"))
        sheet.cell(i, 12, flow.get("trust_net"))
        sheet.cell(i, 13, flow.get("dealer_net"))

        for j, n in enumerate(news[:3]):
            sheet.cell(i, 14 + j, n.get("title"))

    _autofit(sheet, [10, 20, 30, 14, 10, 14, 10, 14, 30, 16, 14, 14, 14, 40, 40, 40])


def generate_excel_report(path: str = EXCEL_REPORT_PATH, holdings=None, ctx=None) -> str:
    if holdings is None:
        holdings = load_portfolio()
        enrich_with_market_data(holdings)

    if ctx is None:
        ctx = build_report_context(holdings)

    wb = Workbook()
    _sheet_overview(wb, holdings, ctx)
    _sheet_fundamentals(wb, holdings)
    _sheet_buy_points(wb, ctx.get("buy_points_data", []))
    _sheet_ai_scoring(wb, holdings, ctx.get("ai_scoring", {}))
    _sheet_sector_allocation(wb, ctx)
    _sheet_risk_rebalance(wb, ctx)
    _sheet_monthly_allocation(wb, ctx.get("monthly_allocation_data", {}))
    _sheet_market_intel(wb, holdings, ctx.get("market_intel", {}))
    _sheet_candidate_watchlist(wb, ctx.get("candidate_watchlist", []))
    _sheet_candidate_watchlist(wb, ctx.get("candidate_watchlist_etf", []), sheet_name="潛力新標的觀察(ETF)")

    disclaimer_sheet = wb.worksheets[-1]
    last_row = disclaimer_sheet.max_row + 2
    disclaimer_sheet.cell(last_row, 1, DISCLAIMER).alignment = Alignment(wrap_text=True)

    _rescale_fonts(wb)
    wb.save(path)
    return path


if __name__ == "__main__":
    saved_path = generate_excel_report()
    print(f"投資報告 Excel 已產生：{saved_path}")
