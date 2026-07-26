import openpyxl

def fill_stock_names(file_path):
    # Mapping of stock codes to names
    name_map = {
        "0050": "元大台灣50",
        "00403A": "統一台股升級50主動式ETF",
        "00685L": "群益臺灣加權正2",
        "00713": "元大台灣高股息低波動",
        "00878": "國泰永續高股息",
        "00918": "大華優利高填息30",
        "00981A": "統一台股增長主動式ETF",
        "00990A": "元大全球AI新經濟主動式ETF",
        "1436": "華友聯",
        "2308": "台達電",
        "2317": "鴻海",
        "2330": "台積電",
        "2454": "聯發科",
        "2886": "兆豐金",
        "3017": "奇鋐",
        "6550": "北極星藥業-KY",
        "009816": "凱基台灣TOP 50 ETF"
    }

    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Find the column indices for '股票編號' and '名稱'
    headers = [cell.value for cell in sheet[1]]
    try:
        code_idx = headers.index("股票編號") + 1
        name_idx = headers.index("名稱") + 1
    except ValueError as e:
        # If columns have different names (e.g. read as garbled in standard display)
        # Let's check headers by position: Col 1 is Code, Col 2 is Name
        code_idx = 1
        name_idx = 2

    # Loop through rows starting from row 2
    for r in range(2, sheet.max_row + 1):
        code_val = str(sheet.cell(r, code_idx).value).strip()
        # Find the name in the map
        if code_val in name_map:
            sheet.cell(r, name_idx).value = name_map[code_val]
            print(f"Filled: {code_val} -> {name_map[code_val]}")
        else:
            print(f"Warning: Stock code {code_val} not in name map")

    wb.save(file_path)
    print("Successfully updated Excel file with stock names.")

if __name__ == "__main__":
    fill_stock_names("個人持股明細.xlsx")
