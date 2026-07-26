"""持股分析核心邏輯：讀取 Excel、抓取即時股價與基本面、產出投資組合分析。"""

import os
from dataclasses import dataclass
from datetime import date
import openpyxl
import yfinance as yf
from dotenv import load_dotenv

load_dotenv()


EXCEL_PATH = "個人持股明細.xlsx"

# 損益率門檻，用於產生加減碼建議
TAKE_PROFIT_THRESHOLD = 0.50       # 獲利超過 50% 才提示考慮分批停利
STOP_LOSS_THRESHOLD = -0.20        # 虧損超過 20% 提示檢視基本面、評估停損或加碼
STOP_LOSS_HARD_THRESHOLD = -0.30   # 虧損擴大到 30%，不論法人看法一律建議停損（例如北極星藥業-KY 一路下滑的情況）
CONCENTRATION_WARNING = 0.30   # 單一持股佔投組市值超過 30% 提示集中度風險（風險評估用，維持不變）
REBALANCE_DRIFT = 0.10         # 市值佔比與成本佔比偏移超過 10 個百分點，提示再平衡

# 每月定期定額的持股配置上限（避免資金一直買同一檔）：個股 25%，大型 ETF 較分散可放寬到 30%
INDIVIDUAL_STOCK_MAX_WEIGHT = 0.25
ETF_MAX_WEIGHT = 0.30

# 分批買進策略：依現價落在哪個價位區間，決定本次投入佔「當月可投資金額」的比例
BUY_LEVEL_1_NORMAL_PCT = 0.20    # Level 1 市場正常（現價高於法人目標價，但未達追高門檻）
BUY_LEVEL_2_FAIR_PCT = 0.30      # Level 2 跌到合理價（現價 ≤ 法人目標價）
BUY_LEVEL_3_BEST_PCT = 0.30      # Level 3 最佳買點（落在近 52 週低點～+30% 區間）
BUY_LEVEL_4_PANIC_PCT = 0.20     # Level 4 恐慌大跌（現價接近或低於 52 週低點）

# 現金流結轉上限：未投入資金最多累積 3 個月的月投入金額，超過就不再無限期等待
MAX_CARRYOVER_MONTHS = 3
CASH_FLOW_SHEET_NAME = "現金流"

# 本月資金最多集中投入前幾名優先順位，不要每檔都買一點（使用者明確要求）
MAX_MONTHLY_FUNDED_CANDIDATES = 3

# 判斷「市場／法人持續看好」的標準：法人評等為買進／強力買進，或法人目標價高於目前股價
BULLISH_RECOMMENDATIONS = {"strong_buy", "buy"}

# MONTHLY_INVESTMENT 是使用者的真實投入金額，不寫死在程式碼裡（比照 GOAL_BASELINE_VALUE 的做法），
# 改從 .env 讀取，.env 本身已 gitignore。
MONTHLY_INVESTMENT = float(os.getenv("MONTHLY_INVESTMENT", "0"))     # 每月定期定額投入金額（新台幣）

# 個人化設定（可投入現金／風險承受能力／投資期限）：這些是使用者的個人設定，不是市場資料，
# 比照 MONTHLY_INVESTMENT 的做法直接寫死在這裡，之後要調整就直接改這幾個值。
AVAILABLE_CASH = 0             # 目前可投入現金（新台幣），供 AI 綜合評分的加碼建議參考
RISK_TOLERANCE = "中等"         # 風險承受能力：保守／中等／積極
INVESTMENT_HORIZON = "中長期"   # 投資期限

# 個人短期目標（2026-07-25 設定，之後不要因為市值變動就跟著改基準）：
# 目標一：一年內累積獲利（總市值-總成本，含新增定期定額投入的部位）達到當初設定基準時的總市值；
# 目標二：期間內任何時候的總市值都不低於基準總市值（下檔保護，不是「賺到目標」才算數，過程中也不能破底）。
# GOAL_BASELINE_VALUE 是使用者的真實總資產數字，不寫死在程式碼裡（會被 commit 進 git 歷史，
# 即使之後改掉也留在歷史紀錄），改從 .env 讀取，.env 本身已 gitignore。
GOAL_BASELINE_DATE = "2026-07-25"
GOAL_TARGET_DATE = "2027-07-25"
GOAL_BASELINE_VALUE = float(os.getenv("GOAL_BASELINE_VALUE", "0"))


@dataclass
class Holding:
    code: str
    name: str
    shares: float
    avg_cost: float
    price: float | None = None
    price_date: str | None = None  # 現價的資料日期，格式 YYYYMMDD（來自 MIS 近即時報價 API 的 d 欄位）
    sector: str | None = None
    industry: str | None = None
    long_name: str | None = None
    trailing_pe: float | None = None
    forward_pe: float | None = None
    trailing_eps: float | None = None
    forward_eps: float | None = None
    target_mean_price: float | None = None
    recommendation: str | None = None
    week52_low: float | None = None
    week52_high: float | None = None
    dividend_yield: float | None = None
    market_cap: float | None = None
    market: str | None = None  # "sii"（上市）或 "otc"（上櫃），依 yfinance 命中的股票代號後綴判斷

    @property
    def cost_value(self) -> float:
        return self.shares * self.avg_cost

    @property
    def market_value(self) -> float | None:
        if self.price is None:
            return None
        return self.shares * self.price

    @property
    def pnl(self) -> float | None:
        if self.market_value is None:
            return None
        return self.market_value - self.cost_value

    @property
    def pnl_pct(self) -> float | None:
        if self.pnl is None or self.cost_value == 0:
            return None
        return self.pnl / self.cost_value

    @property
    def upside_to_target(self) -> float | None:
        """現價相對法人目標價的上漲空間。"""
        if self.target_mean_price is None or self.price is None or self.price == 0:
            return None
        return (self.target_mean_price - self.price) / self.price


def load_portfolio(file_path: str = EXCEL_PATH) -> list[Holding]:
    """從個人持股明細.xlsx 讀取代號、名稱、股數、平均成本。"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    holdings: list[Holding] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code, name, shares, avg_cost = row[0], row[1], row[2], row[3]
        if code is None or shares is None:
            continue
        holdings.append(
            Holding(code=str(code).strip(), name=name or "", shares=float(shares), avg_cost=float(avg_cost or 0))
        )
    return holdings


def load_cash_flow_balance(file_path: str = EXCEL_PATH) -> float:
    """讀取「現金流」分頁（日期／項目／金額／備註），加總「金額」欄得到目前現金池餘額。
    存入／股利／已實現獲利請填正數，已投入／提領請填負數（欄位本身不判斷項目文字，只加總金額）。
    使用者尚未建立這個分頁時（例如舊版 Excel 檔案）優雅地回傳 0，不噴錯。
    """
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        return 0.0
    if CASH_FLOW_SHEET_NAME not in wb.sheetnames:
        return 0.0

    sheet = wb[CASH_FLOW_SHEET_NAME]
    balance = 0.0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        amount = row[2] if len(row) > 2 else None
        if isinstance(amount, (int, float)):
            balance += amount
    return balance


def fetch_market_data(holding: Holding) -> None:
    """依序嘗試 .TW（上市）與 .TWO（上櫃）取得即時股價與基本面資訊。"""
    for suffix in (".TW", ".TWO"):
        ticker = yf.Ticker(f"{holding.code}{suffix}")
        try:
            info = ticker.info
        except Exception:
            continue
        price = info.get("regularMarketPrice") or info.get("previousClose")
        if price:
            holding.price = float(price)
            holding.market = "sii" if suffix == ".TW" else "otc"
            holding.sector = info.get("sector")
            holding.industry = info.get("industry")
            holding.long_name = info.get("longName") or info.get("shortName")
            holding.trailing_pe = info.get("trailingPE")
            holding.forward_pe = info.get("forwardPE")
            holding.trailing_eps = info.get("trailingEps")
            holding.forward_eps = info.get("forwardEps")
            holding.target_mean_price = info.get("targetMeanPrice")
            holding.recommendation = info.get("recommendationKey")
            holding.week52_low = info.get("fiftyTwoWeekLow")
            holding.week52_high = info.get("fiftyTwoWeekHigh")
            holding.dividend_yield = info.get("dividendYield")
            holding.market_cap = info.get("marketCap")
            return


def enrich_with_market_data(holdings: list[Holding]) -> None:
    for h in holdings:
        fetch_market_data(h)

    from market_data import fetch_realtime_quotes

    codes_by_market = {h.code: h.market for h in holdings if h.market}
    quotes = fetch_realtime_quotes(codes_by_market)
    for h in holdings:
        quote = quotes.get(h.code)
        if quote and quote.get("price"):
            h.price = quote["price"]
            h.price_date = quote.get("date")


def format_price_date(raw: str | None) -> str | None:
    """把 MIS API 回傳的 YYYYMMDD 日期字串轉成 `M/D` 顯示格式；格式不對就回傳 None（不硬湊）。"""
    if not raw or len(raw) != 8 or not raw.isdigit():
        return None
    month = int(raw[4:6])
    day = int(raw[6:8])
    return f"{month}/{day}"


def current_price_display(h: "Holding") -> str:
    """「目前股價」欄位顯示格式：`7/24(最後一個收盤日)150.5`；沒有日期資料時只顯示股價，
    完全沒有股價資料時顯示 N/A（缺資料優雅降級，不硬湊）。
    """
    if h.price is None:
        return "N/A"
    price_str = f"{h.price:,.1f}"
    date_str = format_price_date(h.price_date)
    if date_str:
        return f"{date_str}(最後一個收盤日){price_str}"
    return price_str


def total_market_value(holdings: list[Holding]) -> float:
    return sum(h.market_value for h in holdings if h.market_value is not None)


def total_cost_value(holdings: list[Holding]) -> float:
    return sum(h.cost_value for h in holdings)


def weight(holding: Holding, total_value: float) -> float | None:
    if holding.market_value is None or total_value == 0:
        return None
    return holding.market_value / total_value


def cost_weight(holding: Holding, total_cost: float) -> float | None:
    if total_cost == 0:
        return None
    return holding.cost_value / total_cost


# yfinance 回傳的 sector 是它自家（Morningstar 系）分類法，固定 11 種，涵蓋台股常見標的：
SECTOR_NAME_ZH = {
    "Technology": "科技",
    "Financial Services": "金融服務",
    "Healthcare": "醫療保健",
    "Consumer Cyclical": "非必需消費",
    "Consumer Defensive": "必需消費",
    "Industrials": "工業",
    "Basic Materials": "基礎原物料",
    "Energy": "能源",
    "Utilities": "公用事業",
    "Real Estate": "房地產",
    "Communication Services": "通訊服務",
}

# industry 是比 sector 更細的分類，種類遠多於 11 種且沒有官方固定清單，這裡先涵蓋使用者目前
# 真實持股會用到的（Semiconductors／Banks - Regional／Biotechnology／Electronic Components／
# Real Estate - Development 等），加上常見台股類別；查不到的產業原樣顯示英文，不強行硬翻。
INDUSTRY_NAME_ZH = {
    "Semiconductors": "半導體",
    "Semiconductor Equipment & Materials": "半導體設備與材料",
    "Banks - Regional": "區域銀行",
    "Banks - Diversified": "多元銀行",
    "Biotechnology": "生物科技",
    "Electronic Components": "電子零組件",
    "Electronics & Computer Distribution": "電子與電腦通路",
    "Real Estate - Development": "房地產開發",
    "Real Estate Services": "不動產服務",
    "Real Estate - Diversified": "多元不動產",
    "Consumer Electronics": "消費性電子",
    "Computer Hardware": "電腦硬體",
    "Communication Equipment": "通訊設備",
    "Software - Application": "應用軟體",
    "Software - Infrastructure": "基礎架構軟體",
    "Information Technology Services": "資訊科技服務",
    "Specialty Industrial Machinery": "特殊工業機械",
    "Electronic Gaming & Multimedia": "電子遊戲與多媒體",
    "Telecom Services": "電信服務",
    "Insurance - Life": "壽險",
    "Insurance - Property & Casualty": "產物保險",
    "Insurance - Diversified": "多元保險",
    "Insurance Brokers": "保險經紀",
    "Financial Conglomerates": "金融控股",
    "Asset Management": "資產管理",
    "Capital Markets": "資本市場",
    "Credit Services": "信用服務",
    "Drug Manufacturers - General": "製藥（一般）",
    "Drug Manufacturers - Specialty & Generic": "製藥（特用與學名藥）",
    "Medical Devices": "醫療器材",
    "Diagnostics & Research": "檢驗與研發",
    "Medical Instruments & Supplies": "醫療器材與用品",
    "Chemicals": "化學",
    "Specialty Chemicals": "特用化學",
    "Steel": "鋼鐵",
    "Building Materials": "建材",
    "Utilities - Regulated Electric": "電力事業（受監管）",
    "Oil & Gas Refining & Marketing": "油氣煉製與行銷",
    "Airlines": "航空",
    "Marine Shipping": "海運",
    "Railroads": "鐵路運輸",
    "Trucking": "貨運",
    "Integrated Freight & Logistics": "整合物流",
    "Auto Parts": "汽車零組件",
    "Auto Manufacturers": "汽車製造",
    "Household & Personal Products": "家用與個人用品",
    "Packaged Foods": "包裝食品",
    "Beverages - Non-Alcoholic": "非酒精飲料",
    "Restaurants": "餐飲",
    "Apparel Manufacturing": "成衣製造",
    "Textile Manufacturing": "紡織製造",
    "Furnishings, Fixtures & Appliances": "家具家飾與家電",
    "Specialty Retail": "特殊零售",
    "Utilities - Renewable": "再生能源事業",
    "Solar": "太陽能",
}


def sector_label_zh(sector: str | None) -> str | None:
    """sector 顯示用中文名稱；查不到對照就原樣回傳英文（缺資料不亂猜的一貫原則）。"""
    if not sector:
        return None
    return SECTOR_NAME_ZH.get(sector, sector)


def industry_label_zh(industry: str | None) -> str | None:
    """industry 顯示用中文名稱；查不到對照就原樣回傳英文。"""
    if not industry:
        return None
    return INDUSTRY_NAME_ZH.get(industry, industry)


def industry_or_sector_label_zh(h: "Holding") -> str:
    """對應既有 `h.industry or h.sector or "N/A"` pattern 的中文版，四個輸出檔案共用。"""
    return industry_label_zh(h.industry) or sector_label_zh(h.sector) or "N/A"


def risk_assessment(holdings: list[Holding]) -> list[str]:
    """持股集中度風險評估。"""
    total_value = total_market_value(holdings)
    warnings: list[str] = []
    if total_value == 0:
        return ["無法取得市值資料，無法評估風險。"]

    for h in holdings:
        w = weight(h, total_value)
        if w is not None and w >= CONCENTRATION_WARNING:
            warnings.append(f"{h.code} {h.name}：佔投組市值 {w:.1%}，集中度偏高，建議留意單一標的風險。")

    sector_totals: dict[str, float] = {}
    for h in holdings:
        if h.sector and h.market_value is not None:
            sector_totals[h.sector] = sector_totals.get(h.sector, 0) + h.market_value
    for sector, value in sector_totals.items():
        w = value / total_value
        if w >= CONCENTRATION_WARNING:
            warnings.append(f"產業「{sector_label_zh(sector)}」佔投組市值 {w:.1%}，產業集中度偏高。")

    if not warnings:
        warnings.append("目前無明顯集中度風險。")
    return warnings


def is_market_bullish(h: Holding) -> bool:
    """判斷「市場／法人持續看好」：法人評等為買進／強力買進，或法人目標價高於目前股價。"""
    if h.recommendation in BULLISH_RECOMMENDATIONS:
        return True
    if h.upside_to_target is not None and h.upside_to_target > 0:
        return True
    return False


def suggestions(holdings: list[Holding]) -> list[str]:
    """依損益率的簡易規則型加減碼建議（非專業投資建議，僅供參考）。

    獲利達 +50% 或虧損達 -20% 時，會再參考法人評等／目標價判斷市場是否持續看好：
    看好則建議續抱（並附上法人目標價），不盲目停利出清或停損；
    虧損擴大到 -30% 則不論法人看法，一律建議停損。
    """
    result = []
    for h in holdings:
        if h.pnl_pct is None:
            result.append(f"{h.code} {h.name}：尚無市價資料，無法給出建議。")
            continue

        bullish = is_market_bullish(h)
        target_note = f"，法人目標價 {h.target_mean_price:,.1f} 元" if bullish and h.target_mean_price is not None else ""

        if h.pnl_pct <= STOP_LOSS_HARD_THRESHOLD:
            result.append(f"{h.code} {h.name}：虧損擴大至 {h.pnl_pct:.1%}，不論法人看法，建議停損。")
        elif h.pnl_pct >= TAKE_PROFIT_THRESHOLD:
            if bullish:
                result.append(f"{h.code} {h.name}：獲利 {h.pnl_pct:.1%}，但法人／市場持續看好{target_note}，建議續抱，暫不停利出清。")
            else:
                result.append(f"{h.code} {h.name}：獲利 {h.pnl_pct:.1%}，可考慮分批停利。")
        elif h.pnl_pct <= STOP_LOSS_THRESHOLD:
            if bullish:
                result.append(f"{h.code} {h.name}：虧損 {h.pnl_pct:.1%}，但法人／市場持續看好{target_note}，建議續抱觀察。")
            else:
                result.append(f"{h.code} {h.name}：虧損 {h.pnl_pct:.1%}，建議檢視基本面，評估停損或逢低加碼。")
        else:
            result.append(f"{h.code} {h.name}：損益 {h.pnl_pct:.1%}，建議持續持有觀察。")
    return result


def reasonable_buy_point(h: Holding) -> str:
    """合理買點：綜合近 52 週區間與法人目標價估算。"""
    if h.week52_low is None or h.week52_high is None:
        return f"{h.code} {h.name}：尚無近一年價格區間資料，暫無法估算買點。"

    buy_zone_low = h.week52_low
    buy_zone_high = h.week52_low + (h.week52_high - h.week52_low) * 0.3

    parts = [f"近一年區間 {h.week52_low:,.1f}～{h.week52_high:,.1f}，逢低承接參考區間約 {buy_zone_low:,.1f}～{buy_zone_high:,.1f}"]

    if h.price is not None:
        if h.price <= buy_zone_high:
            parts.append("現價已落在偏低區間，屬於合理承接位置")
        else:
            parts.append("現價高於偏低承接區間，建議等待拉回或分批布局")

    if h.upside_to_target is not None:
        direction = "具上漲空間" if h.upside_to_target > 0 else "已接近或超過法人目標價"
        parts.append(f"法人平均目標價 {h.target_mean_price:,.1f}，{direction}（{h.upside_to_target:+.1%}）")

    return f"{h.code} {h.name}：" + "；".join(parts) + "。"


def buy_point_data(h: Holding) -> dict:
    """`reasonable_buy_point()` 的結構化平行版本，供表格渲染用（`reasonable_buy_point()` 本身
    保留不動，怕有其他地方依賴那個句子字串）。缺近一年價格區間資料時，區間／建議承接區間欄位
    給 None，優雅降級。"""
    data = {
        "code": h.code,
        "name": h.name,
        "week52_range": None,
        "suggested_zone": None,
        "price_position": None,
        "target_price": h.target_mean_price,
        "upside_pct": h.upside_to_target,
    }
    if h.week52_low is None or h.week52_high is None:
        data["price_position"] = "尚無近一年價格區間資料，暫無法估算買點。"
        return data

    buy_zone_low = h.week52_low
    buy_zone_high = h.week52_low + (h.week52_high - h.week52_low) * 0.3
    data["week52_range"] = f"{h.week52_low:,.1f}～{h.week52_high:,.1f}"
    data["suggested_zone"] = f"{buy_zone_low:,.1f}～{buy_zone_high:,.1f}"

    if h.price is not None:
        if h.price <= buy_zone_high:
            data["price_position"] = "現價已落在偏低區間，屬於合理承接位置"
        else:
            data["price_position"] = "現價高於偏低承接區間，建議等待拉回或分批布局"

    return data


def rebalance_check(holdings: list[Holding]) -> list[str]:
    """比較市值佔比與成本佔比的偏移，評估是否需要再平衡。"""
    total_value = total_market_value(holdings)
    total_cost = total_cost_value(holdings)
    if total_value == 0 or total_cost == 0:
        return ["尚無足夠資料評估再平衡需求。"]

    messages = []
    for h in holdings:
        w = weight(h, total_value)
        cw = cost_weight(h, total_cost)
        if w is None or cw is None:
            continue
        drift = w - cw
        if drift >= REBALANCE_DRIFT:
            messages.append(f"{h.code} {h.name}：市值佔比較原始成本佔比高出 {drift:.1%}，漲多墊高集中度，建議部分獲利了結以回歸配置。")
        elif drift <= -REBALANCE_DRIFT:
            messages.append(f"{h.code} {h.name}：市值佔比較原始成本佔比低了 {abs(drift):.1%}，若基本面仍佳可考慮加碼回補。")

    if not messages:
        messages.append("目前配置與原始成本比重偏移不大，暫不需要再平衡。")
    return messages


def sector_allocation(holdings: list[Holding]) -> dict[str, float]:
    """各產業佔投組市值比重（回傳的 key 已是中文顯示名稱，供報告直接顯示；沒有其他函式依賴
    這個回傳值的 key 是英文，可以直接在回傳時翻譯）。"""
    total_value = total_market_value(holdings)
    result: dict[str, float] = {}
    if total_value == 0:
        return result
    for h in holdings:
        key = sector_label_zh(h.sector) or ETF_SECTOR_LABEL
        if h.market_value is not None:
            result[key] = result.get(key, 0) + h.market_value / total_value
    return dict(sorted(result.items(), key=lambda x: x[1], reverse=True))


ETF_SECTOR_LABEL = "未分類（ETF／其他）"


def sector_allocation_advice(holdings: list[Holding]) -> list[str]:
    allocation = sector_allocation(holdings)
    advice = []
    for sector, w in allocation.items():
        if sector == ETF_SECTOR_LABEL:
            continue  # ETF 本身已是分散持股，不視為單一產業集中風險
        if w >= CONCENTRATION_WARNING:
            advice.append(f"「{sector}」佔比 {w:.1%}，偏重單一產業，建議逐步分散至其他產業（如金融、內需、原物料）以降低景氣循環風險。")
    if not advice:
        advice.append("個股部位的產業配置尚稱分散，暫無明顯調整必要。")
    return advice


def _buy_level_tier(h: Holding, fair_value: float | None) -> tuple[str, float, str]:
    """依現價相對合理價／52 週區間，判斷落在哪個分批買進 Level，並附上「等待成本」提示
    （比較法人預估的上漲空間 vs 目前的高估幅度，決定要提前布局還是等待拉回）。
    """
    price = h.price
    if price is None:
        return "Level 1：市場正常", BUY_LEVEL_1_NORMAL_PCT, ""

    if h.week52_low is not None and price <= h.week52_low * 1.05:
        return "Level 4：恐慌大跌", BUY_LEVEL_4_PANIC_PCT, ""

    if h.week52_low is not None and h.week52_high is not None:
        buy_zone_high = h.week52_low + (h.week52_high - h.week52_low) * 0.3
        if price <= buy_zone_high:
            return "Level 3：最佳買點", BUY_LEVEL_3_BEST_PCT, ""

    if fair_value is not None and price <= fair_value:
        return "Level 2：跌到合理價", BUY_LEVEL_2_FAIR_PCT, ""

    note = ""
    if fair_value is not None and price > fair_value:
        overvaluation = (price - fair_value) / fair_value
        upside = h.upside_to_target
        if upside is not None and upside - overvaluation >= 0.05:
            note = f"（現價略高於合理價 {overvaluation:.1%}，但法人預估仍有 {upside:+.1%} 上漲空間，等待成本較高，可考慮提前布局）"
        elif overvaluation >= 0.20:
            note = f"（現價已高於合理價 {overvaluation:.1%}，等待成本較低，建議等待拉回）"
    return "Level 1：市場正常", BUY_LEVEL_1_NORMAL_PCT, note


def compute_monthly_allocation(
    holdings: list[Holding],
    ai_scoring: dict | None = None,
    cash_flow_balance: float = 0.0,
    amount: float = MONTHLY_INVESTMENT,
) -> dict:
    """`monthly_allocation_suggestion()` 的結構化版本：算出優先購買清單、本月資金分配明細等
    結構化資料（供表格渲染用），沒有觸發的情境對應欄位給 None。計算邏輯與
    `monthly_allocation_suggestion()`（現在改為呼叫本函式再格式化成文字）完全共用，
    不會有兩份計算邏輯逐漸失去同步的風險。
    """
    ai_scoring = ai_scoring or {}
    total_value = total_market_value(holdings)
    result: dict = {
        "total_budget": None,
        "amount": amount,
        "carryover": None,
        "carryover_cap": None,
        "priority_list": [],
        "allocations": [],
        "remaining": None,
        "excluded_notes": [],
        "no_candidates_note": None,
        "no_allocation_note": None,
        "no_value_note": None,
    }
    if total_value == 0:
        result["no_value_note"] = "尚無市值資料，無法規劃本月投入建議。"
        return result

    carryover_cap = amount * MAX_CARRYOVER_MONTHS
    carryover = max(0.0, min(cash_flow_balance, carryover_cap))
    total_budget = amount + carryover
    result["carryover_cap"] = carryover_cap
    result["carryover"] = carryover
    result["total_budget"] = total_budget

    candidates = []
    excluded_notes = []
    for h in holdings:
        w = weight(h, total_value)
        max_weight = ETF_MAX_WEIGHT if not h.sector else INDIVIDUAL_STOCK_MAX_WEIGHT
        if w is None:
            continue
        if w >= max_weight:
            excluded_notes.append(f"{h.code} {h.name} 佔投組已達 {w:.1%}（上限 {max_weight:.0%}），本月不建議再加碼，避免過度集中單一標的。")
            continue
        if h.pnl_pct is not None and h.pnl_pct <= STOP_LOSS_THRESHOLD:
            excluded_notes.append(f"{h.code} {h.name} 已列入停損觀察名單（虧損 {h.pnl_pct:.1%}），本月新資金不自動加碼，請先確認基本面是否轉壞。")
            continue
        r = ai_scoring.get(h.code) or {}
        if r.get("signal") == "🔴" and "追價" in (r.get("label") or ""):
            excluded_notes.append(f"{h.code} {h.name}：現價已追高，本月不建議加碼，等待拉回。")
            continue
        candidates.append(h)

    result["excluded_notes"] = excluded_notes

    if not candidates:
        result["no_candidates_note"] = "所有持股皆已達配置上限或處於停損觀察／追高禁區，本月建議暫緩加碼，資金結轉下月。"
        return result

    def _priority_key(h: Holding) -> float:
        r = ai_scoring.get(h.code) or {}
        inv = r["investment_score"] if r.get("investment_score") is not None else 0.0
        conv = r["conviction_score"] if r.get("conviction_score") is not None else 0.0
        return inv + conv * 0.2  # Investment Score 為主，Conviction Score 當作次要加分

    candidates.sort(key=_priority_key, reverse=True)

    priority_list = []
    for i, h in enumerate(candidates[:5], start=1):
        r = ai_scoring.get(h.code) or {}
        priority_list.append(
            {
                "rank": i,
                "code": h.code,
                "name": h.name,
                "investment_score": r.get("investment_score"),
                "conviction_stars": r.get("conviction_stars", "N/A"),
            }
        )
    result["priority_list"] = priority_list

    remaining = total_budget
    allocated_any = False
    min_meaningful = total_budget * 0.05  # 低於本月預算 5% 的零碎金額不再往下一順位分配，直接結轉下月
    allocations = []
    for h in candidates[:MAX_MONTHLY_FUNDED_CANDIDATES]:
        if remaining < min_meaningful:
            break
        r = ai_scoring.get(h.code) or {}
        fair_value = r.get("fair_value")
        tier_label, tier_pct, note = _buy_level_tier(h, fair_value)
        alloc_amount = total_budget * tier_pct  # 每個優先順位的投入比例都以「本月總預算」為基準，不是剩餘金額的比例，避免越後面越零碎
        alloc_amount = min(alloc_amount, remaining)
        if alloc_amount < min_meaningful or not h.price:
            continue
        allocations.append(
            {
                "code": h.code,
                "name": h.name,
                "amount": alloc_amount,
                "tier_label": tier_label,
                "tier_pct": tier_pct,
                "shares": alloc_amount / h.price,
                "note": note,
            }
        )
        remaining -= alloc_amount
        allocated_any = True

    result["allocations"] = allocations
    result["remaining"] = remaining if remaining >= min_meaningful else None
    if not allocated_any:
        result["no_allocation_note"] = "優先候選標的現價目前都偏高，本月建議暫緩投入，資金結轉下月。"

    return result


def _format_monthly_allocation(data: dict) -> list[str]:
    """把 `compute_monthly_allocation()` 的結構化資料組回跟舊版 `monthly_allocation_suggestion()`
    完全相同的文字輸出（daily_report.py／excel_report.py 既有的文字消費端不能改變其輸出）。
    """
    if data.get("no_value_note"):
        return [data["no_value_note"]]

    amount = data["amount"]
    carryover = data["carryover"]
    total_budget = data["total_budget"]
    result = [
        f"本月可投資金額：新台幣 {total_budget:,.0f} 元"
        + (f"（月定期定額 {amount:,.0f} 元＋現金流結餘 {carryover:,.0f} 元，結轉上限 {data['carryover_cap']:,.0f} 元）" if carryover else f"（月定期定額 {amount:,.0f} 元）")
    ]

    if data.get("no_candidates_note"):
        result.append(data["no_candidates_note"])
        result.extend(data["excluded_notes"])
        return result

    result.append("")
    result.append("優先購買清單（依 Investment Score／長期持股信心排序）：")
    for p in data["priority_list"]:
        inv_str = f"{p['investment_score']:.0f}" if p.get("investment_score") is not None else "N/A"
        result.append(f"{p['rank']}. {p['code']} {p['name']}：Investment Score {inv_str}／長期信心 {p.get('conviction_stars', 'N/A')}")

    result.append("")
    result.append("本月資金分配（集中在優先順序最前面的標的，不是每檔都買一點；資金用完或達上限為止）：")
    for a in data["allocations"]:
        result.append(
            f"- {a['code']} {a['name']}：投入 {a['amount']:,.0f} 元（{a['tier_label']}，本輪比例 {a['tier_pct']:.0%}；"
            f"約可買 {a['shares']:,.1f} 股）{a['note']}"
        )

    if data["remaining"] is not None:
        result.append(f"尚未分配：{data['remaining']:,.0f} 元，將結轉至下月（結轉上限：{MAX_CARRYOVER_MONTHS} 個月投入金額 {data['carryover_cap']:,.0f} 元）。")
    if data.get("no_allocation_note"):
        result.append(data["no_allocation_note"])

    result.extend(data["excluded_notes"])
    return result


def monthly_allocation_suggestion(
    holdings: list[Holding],
    ai_scoring: dict | None = None,
    cash_flow_balance: float = 0.0,
    amount: float = MONTHLY_INVESTMENT,
) -> list[str]:
    """依 AI 綜合評分排出優先購買清單，把本月可投資金額集中投入排名最前面的標的
    （不是每檔都買一點），並依現價落在哪個價位區間決定分批投入比例（見 `_buy_level_tier()`）。

    薪資現金流投資人的現金流優先原則：本月可投資金額 = 每月固定投入 + 現金流結餘
    （現金流結餘＝股利/已實現獲利/手動存入累計，來自「現金流」Excel 分頁，結轉上限
    為 `MAX_CARRYOVER_MONTHS` 個月的投入金額，避免無限期等待理想買點而長期空手）。
    """
    data = compute_monthly_allocation(holdings, ai_scoring, cash_flow_balance, amount)
    return _format_monthly_allocation(data)


def market_intel_highlight(info: dict | None) -> str:
    """從單一持股的 market_intel 資料中，依優先序（重大訊息 > 三大法人買賣超 > 月營收年增率 >
    下一場法說會）只挑一項最值得注意的重點，組成一句話（給「重點摘要」欄位用）。
    全部沒資料時回傳「暫無特別重點」，不硬湊。
    """
    if not info:
        return "暫無特別重點"

    material_news = info.get("material_news") or []
    if material_news:
        subject = (material_news[0].get("主旨") or "").strip()
        if subject:
            return f"重大訊息：{subject}"

    flow = info.get("institutional_flow")
    if flow and flow.get("total_net") is not None:
        if flow.get("foreign_net") is not None:
            return (
                f"三大法人買賣超 {flow['total_net']:+,.0f} 股"
                f"（外資 {flow['foreign_net']:+,.0f}／投信 {flow['trust_net']:+,.0f}／自營商 {flow['dealer_net']:+,.0f}）"
            )
        return f"三大法人買賣超合計 {flow['total_net']:+,.0f} 股"

    rev = info.get("monthly_revenue")
    if rev:
        yoy = rev.get("營業收入-去年同月增減(%)")
        month = rev.get("資料年月")
        if yoy not in (None, ""):
            return f"月營收({month})年增率 {float(yoy):+.1f}%"

    conf = info.get("investor_conference")
    if conf:
        return f"下一場法說會 {conf['date']} {conf['time']}，地點：{conf['location']}"

    return "暫無特別重點"


def goal_progress(total_value: float, total_pnl: float) -> dict:
    """個人短期目標追蹤：一年內累積獲利達基準總市值、且過程中總市值不低於基準總市值。
    `total_pnl`（總市值-總成本）用累積成本當分母，新增的定期定額投入不會虛增這個數字，
    是這裡拿來對照獲利目標的正確指標；`total_value` 的絕對水位則用來檢查有沒有跌破下限。
    """
    today = date.today()
    target_date = date.fromisoformat(GOAL_TARGET_DATE)
    baseline_date = date.fromisoformat(GOAL_BASELINE_DATE)

    total_days = (target_date - baseline_date).days
    days_elapsed = (today - baseline_date).days
    days_remaining = (target_date - today).days

    profit_progress_pct = total_pnl / GOAL_BASELINE_VALUE if GOAL_BASELINE_VALUE else None
    floor_breached = total_value < GOAL_BASELINE_VALUE

    expected_pnl_by_now = GOAL_BASELINE_VALUE * (days_elapsed / total_days) if total_days else None
    on_pace = (total_pnl >= expected_pnl_by_now) if expected_pnl_by_now is not None else None

    return {
        "baseline_date": GOAL_BASELINE_DATE,
        "target_date": GOAL_TARGET_DATE,
        "baseline_value": GOAL_BASELINE_VALUE,
        "profit_target": GOAL_BASELINE_VALUE,
        "current_pnl": total_pnl,
        "current_value": total_value,
        "profit_progress_pct": profit_progress_pct,
        "floor_breached": floor_breached,
        "days_remaining": days_remaining,
        "on_pace": on_pace,
    }


def build_report_context(holdings: list[Holding], include_market_intel: bool = True) -> dict:
    """整合所有分析結果，供報告輸出使用。"""
    total_value = total_market_value(holdings)
    total_cost = total_cost_value(holdings)
    total_pnl = total_value - total_cost
    total_pnl_pct = (total_pnl / total_cost) if total_cost else None

    market_intel = {}
    market_index = None
    ai_scoring = {}
    candidate_watchlist = []
    candidate_watchlist_etf = []
    if include_market_intel:
        from market_data import build_market_intel, fetch_market_index
        from technicals import build_technicals_map
        from market_environment import fetch_market_environment, compute_environment_score
        from scoring import build_scoring_context
        from candidate_watchlist import evaluate_candidates, evaluate_etf_candidates

        market_intel = build_market_intel(holdings)
        market_index = fetch_market_index()
        technicals_map = build_technicals_map(holdings)
        environment_score = compute_environment_score(fetch_market_environment())
        ai_scoring = build_scoring_context(holdings, market_intel, technicals_map, environment_score)
        candidate_watchlist = evaluate_candidates(holdings, environment_score)
        candidate_watchlist_etf = evaluate_etf_candidates(holdings, environment_score)

    cash_flow_balance = load_cash_flow_balance()
    monthly_allocation_data = compute_monthly_allocation(holdings, ai_scoring, cash_flow_balance)

    return {
        "holdings": holdings,
        "total_value": total_value,
        "total_cost": total_cost,
        "total_pnl": total_pnl,
        "total_pnl_pct": total_pnl_pct,
        "goal": goal_progress(total_value, total_pnl),
        "risk": risk_assessment(holdings),
        "suggestions": suggestions(holdings),
        "buy_points": [reasonable_buy_point(h) for h in holdings],
        "buy_points_data": [buy_point_data(h) for h in holdings],
        "rebalance": rebalance_check(holdings),
        "sector_allocation": sector_allocation(holdings),
        "sector_advice": sector_allocation_advice(holdings),
        "monthly_allocation": _format_monthly_allocation(monthly_allocation_data),
        "monthly_allocation_data": monthly_allocation_data,
        "monthly_focus": MONTHLY_FOCUS,
        "market_intel": market_intel,
        "market_index": market_index,
        "ai_scoring": ai_scoring,
        "available_cash": AVAILABLE_CASH,
        "risk_tolerance": RISK_TOLERANCE,
        "investment_horizon": INVESTMENT_HORIZON,
        "cash_flow_balance": cash_flow_balance,
        "candidate_watchlist": candidate_watchlist,
        "candidate_watchlist_etf": candidate_watchlist_etf,
    }


MONTHLY_FOCUS = [
    "檢視本月投組績效與大盤（加權指數）的相對表現",
    "檢查各持股是否有重大新聞、法說會或月營收公告",
    "確認持股與產業集中度是否超過警戒線（單一標的／產業 30%）",
    "檢視是否有觸及停利／停損條件的標的",
    "依估值吸引力與再平衡需求，決定本月 10,000 元定期定額的投入標的",
]
